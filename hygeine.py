import pandas as pd
from openpyxl import load_workbook
import difflib

def fill_hygiene_sheet(master_data, format_file_path, hygiene_input_file_path):
    try:
        # File paths
        input_file = hygiene_input_file_path
        output_file = format_file_path
        output_file_filled = format_file_path

        # Load the output Excel file and select the Hygiene sheet
        wb = load_workbook(output_file)
        ws = wb['Annexure- Hygiene Obs']

        # Fill header data (Date and PSV Quarter)
        print("Filling header data...")
        ws['C2'] = master_data.get("Date of Audit")
        ws['C3'] = master_data.get("PSV Quarter")

        # Debug print to see what data we have
        print("\nMaster Data Contents:")
        for key, value in master_data.items():
            print(f"{key}: {value}")

        # Read input Excel file
        input_df = pd.read_excel(input_file)
        input_columns = list(input_df.columns)
        input_columns_lower = [col.lower() for col in input_columns]

        # Get S Loc Code and Category from master_data
        s_loc_code = master_data.get("S Loc Code")
        category = master_data.get("Category")

        print(f"\nRetrieved values:")
        print(f"S Loc Code: {s_loc_code}")
        print(f"Category: {category}")

        if not s_loc_code or not category:
            print("S Loc Code or Category not found in master data.")
            print(f"Available data: {master_data}")  # Debug print
            return

        # Convert input DataFrame's Product column to string and lowercase for comparison
        input_df['Category'] = input_df['Category'].astype(str).str.strip().str.lower()
        
        # Find the row with the matching S Loc Code and Category (case-insensitive)
        row = input_df[
            (input_df['S Loc Code'] == int(s_loc_code)) & 
            (input_df['Category'] == category.strip().lower())
        ]
        
        if row.empty:
            print(f'No data found for S Loc Code {s_loc_code} and Category {category} in input file.')
            # Debug print to show available categories
            print("\nAvailable categories in input file:")
            print(input_df['Category'].unique())
            return
        row = row.iloc[0]

        # Dynamically find the header row containing 'Check Points' and 'Auditor's response'
        header_row = None
        check_points_col = None
        auditor_response_col = None
        for i in range(1, ws.max_row + 1):
            row_values = [ws.cell(row=i, column=col).value for col in range(1, ws.max_column + 1)]
            if 'Check Points' in row_values and "Auditor's response" in row_values:
                header_row = i
                check_points_col = row_values.index('Check Points') + 1
                auditor_response_col = row_values.index("Auditor's response") + 1
                break
        if header_row is None:
            print('Could not find required columns in Hygiene sheet.')
            return

        # Iterate through Check Points and fill Auditor's response
        row_idx = header_row + 1
        while True:
            check_point = ws.cell(row=row_idx, column=check_points_col).value
            if check_point is None:
                break
            # Try exact (case-insensitive) match first
            match_col = None
            for idx, col in enumerate(input_columns_lower):
                if check_point and check_point.lower().strip() == col.strip():
                    match_col = input_columns[idx]
                    break
            # If not found, use fuzzy matching
            if not match_col and check_point:
                close_matches = difflib.get_close_matches(check_point.lower().strip(), input_columns_lower, n=1, cutoff=0.7)
                if close_matches:
                    match_col = input_columns[input_columns_lower.index(close_matches[0])]
            # Fill value if match found
            if match_col:
                ws.cell(row=row_idx, column=auditor_response_col).value = row[match_col]
            row_idx += 1

        # Save the filled output file
        wb.save(output_file_filled)
        print(f'Successfully filled output saved as {output_file_filled}')

    except Exception as e:
        print(f"Error processing hygiene sheet: {str(e)}")
    finally:
        if 'wb' in locals():
            wb.close()

if __name__ == "__main__":
    fill_hygiene_sheet()
