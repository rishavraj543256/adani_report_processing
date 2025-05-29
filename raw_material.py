import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, Protection
from copy import copy

def process_mb52_stock(excel_path):
    # Load the workbook with data_only=True to get calculated values
    wb = load_workbook(filename=excel_path, data_only=True)
    sheet = wb['Mb52- Stock Report']
    
    # Read the regular data with pandas
    mb52_df = pd.read_excel(
        excel_path, 
        sheet_name='Mb52- Stock Report', 
        header=1
    )
    
    # Get the column index for Unrestricted
    header_row = 2  # since header=1 in pandas
    unrestricted_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=header_row, column=col).value == 'Unrestricted':
            unrestricted_col = col
            break
    
    # Get calculated values for Unrestricted
    if unrestricted_col:
        unrestricted_values = []
        for row in range(header_row + 1, sheet.max_row + 1):
            value = sheet.cell(row=row, column=unrestricted_col).value
            if isinstance(value, (int, float)):  # Only append numeric values
                unrestricted_values.append(value)
            else:
                unrestricted_values.append('')
        
        # Replace the Unrestricted column in mb52_df
        mb52_df['Unrestricted'] = pd.Series(unrestricted_values[:len(mb52_df)])
    
    # Create a DataFrame with the exact structure needed - no grouping
    result_df = pd.DataFrame({
        'S Loc Code': mb52_df['Storage Location'],
        'Material Code': mb52_df['Material'],
        'Material Name': mb52_df['Material Description'],
        'UOM IN MT / No.': mb52_df['Base Unit of Measure'],
        'Closing Balance - Net Weight (SAP)': mb52_df['Unrestricted'].fillna(''),
        'Physical Stock - Net Weight': '',
        'Stock under Fumigation': '',
        #'Total Physical Stock': '',
        # 'Actual Shortage / Excess (+/-)': '',
        # 'Remark 1': '',
        # 'General Remark': ''
    })
    
    # Filter out rows where S Loc Code or Material Code is NaN
    result_df = result_df.dropna(subset=['S Loc Code', 'Material Code'])
    
    # Sort the data if needed
    result_df = result_df.sort_values(['S Loc Code', 'Material Code'])
    
    # Filter out any existing "Total" rows from the source data
    result_df = result_df[result_df['Material Name'] != 'Total']
    
    # Calculate total for Closing Balance column
    total = pd.to_numeric(result_df['Closing Balance - Net Weight (SAP)'], errors='coerce').sum()
    
    # Add Total row
    total_row = pd.DataFrame({
        'S Loc Code': '',
        'Material Code': '',
        'Material Name': 'Total',
        'UOM IN MT / No.': '',
        'Closing Balance - Net Weight (SAP)': total,
        'Physical Stock - Net Weight': '',
        'Stock under Fumigation': '',
        #'Total Physical Stock': '',
        # 'Actual Shortage / Excess (+/-)': '',
        # 'Remark 1': '',
        # 'General Remark': ''
    }, index=[0])
    
    # Concatenate the result_df with total_row
    result_df = pd.concat([result_df, total_row], ignore_index=True)
    
    return result_df

def copy_cell_format(source_cell, target_cell):
    """Copy all formatting from source cell to target cell"""
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)

def find_signoff_section(sheet):
    """Find the start and end row of sign-off section"""
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and "S Loc Incharge" in str(cell_value):
                # Find the end of sign-off section
                end_row = row
                while end_row <= sheet.max_row:
                    if all(sheet.cell(row=end_row, column=col).value is None 
                          for col in range(1, sheet.max_column + 1)):
                        end_row -= 1
                        break
                    end_row += 1
                return row, end_row
    return None, None

def move_range_down(sheet, start_row, end_row, shift_amount):
    """Move a range of cells down by shift_amount rows"""
    if start_row is None or end_row is None:
        return
        
    # Move from bottom to top to avoid overwriting
    for row in range(end_row, start_row - 1, -1):
        for col in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=row, column=col)
            target_cell = sheet.cell(row=row + shift_amount, column=col)
            
            # Copy value and style
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            
            # Clear source cell
            source_cell.value = None

def find_header_row(sheet):
    """Find the row containing the column headers"""
    for row in range(1, sheet.max_row + 1):
        cell_value = str(sheet.cell(row=row, column=2).value or '').strip()  # Look in column B
        if cell_value == 'S Loc Code':
            return row
    return None

def find_total_row(sheet, header_row, sign_off_start):
    """Find the row containing 'Total' label"""
    for row in range(header_row + 1, sign_off_start):
        cell_value = str(sheet.cell(row=row, column=4).value or '').strip()  # Column D (Material Name)
        if cell_value == 'Total':
            return row
    return None

def process_rm_stack_wise(excel_path):
    """
    Process the RM Stack Wise sheet to get physical stock and fumigation stock data
    Returns two dictionaries mapping material codes to their respective net weights
    """
    print("\n=== Processing RM Stack Wise Sheet ===")
    # Read the RM Stack Wise sheet
    print(f"Reading Excel file: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name='RM- Stack wise', header=6)
    print(f"Total rows read: {len(df)}")

    # Find the index where 'Stock under Fumigation' appears
    fumigation_start_idx = None
    for i, row in df.iterrows():
        if isinstance(row.get('Stack No. (With Stock)'), str) and 'Stock under Fumigation' in row.get('Stack No. (With Stock)', ''):
            fumigation_start_idx = i
            break

    if fumigation_start_idx is not None:
        general_data = df.iloc[:fumigation_start_idx]
        fumigation_data = df.iloc[fumigation_start_idx+1:]
    else:
        general_data = df
        fumigation_data = pd.DataFrame()  # empty

    # Clean up general_data and fumigation_data
    general_data = general_data[general_data['Stack No. (With Stock)'].notna()]
    general_data = general_data[general_data['Stack No. (With Stock)'] != 'Total Stock']

    fumigation_data = fumigation_data[fumigation_data['Stack No. (With Stock)'].notna()]
    fumigation_data = fumigation_data[fumigation_data['Stack No. (With Stock)'] != 'Total Stock']

    # Initialize dictionaries to store results
    physical_stock = {}
    fumigation_stock = {}

    # Process general section
    for _, row in general_data.iterrows():
        material_code = row['Material Code']
        if pd.notna(material_code):
            try:
                qty_in_mt = pd.to_numeric(row.get('Qty. In MT', 0), errors='coerce')
                bardana_weight = pd.to_numeric(row.get('Bardana Weight', 0), errors='coerce')
                if pd.notna(qty_in_mt) and pd.notna(bardana_weight):
                    net_weight = float(qty_in_mt) - float(bardana_weight)
                    physical_stock[material_code] = physical_stock.get(material_code, 0) + net_weight
            except (ValueError, TypeError):
                continue

    # Process fumigation section
    for _, row in fumigation_data.iterrows():
        material_code = row['Material Code']
        if pd.notna(material_code):
            try:
                qty_in_mt = pd.to_numeric(row.get('Qty. In MT', 0), errors='coerce')
                bardana_weight = pd.to_numeric(row.get('Bardana Weight', 0), errors='coerce')
                if pd.notna(qty_in_mt) and pd.notna(bardana_weight):
                    net_weight = float(qty_in_mt) - float(bardana_weight)
                    fumigation_stock[material_code] = fumigation_stock.get(material_code, 0) + net_weight
            except (ValueError, TypeError):
                continue

    print("\nFinal Stock Summary:")
    print("Physical Stock:", physical_stock)
    print("Fumigation Stock:", fumigation_stock)
    return physical_stock, fumigation_stock

def update_annexure_sheet(format_file_path, processed_data, master_data):
    print("\n=== Updating Annexure Sheet ===")
    # Get physical stock and fumigation stock data
    print("Getting physical and fumigation stock data...")
    physical_stock_dict, fumigation_stock_dict = process_rm_stack_wise(format_file_path)
    
    # Load the workbook
    print(f"Loading workbook: {format_file_path}")
    wb = load_workbook(format_file_path)
    annexure_sheet = wb['Annexure- Raw Material']

    # Fill header data (Date and PSV Quarter)
    print("Filling header data...")
    annexure_sheet['C2'] = master_data.get("Date of Audit")
    annexure_sheet['C3'] = master_data.get("PSV Quarter")
    
    # Find the sign-off section
    print("Looking for sign-off section...")
    sign_off_start, sign_off_end = find_signoff_section(annexure_sheet)
    
    if sign_off_start is None:
        raise ValueError("Could not find sign-off section (S Loc Incharge row) in the Annexure sheet")
    print(f"Found sign-off section: rows {sign_off_start} to {sign_off_end}")
    
    # Find the header row
    print("Looking for header row...")
    header_row = find_header_row(annexure_sheet)
    
    if header_row is None:
        raise ValueError("Could not find the header row with 'S Loc Code' in column B")
    print(f"Found header row: {header_row}")
    
    data_start_row = header_row + 1
    print(f"Data will start at row: {data_start_row}")
    
    # Find the first existing Total row
    total_row_num = None
    for row in range(data_start_row, sign_off_start):
        if str(annexure_sheet.cell(row=row, column=4).value or '').strip() == 'Total':
            total_row_num = row
            print(f"Found existing Total row at: {total_row_num}")
            break
    
    # Store the Total row formatting (from first found Total row)
    total_row_info = None
    if total_row_num:
        total_row_info = []
        for col in range(1, annexure_sheet.max_column + 1):
            cell = annexure_sheet.cell(row=total_row_num, column=col)
            total_row_info.append((None, cell._style if cell.has_style else None))
    
    # Calculate required rows and shift amount
    required_rows = len(processed_data)  # We'll add the Total row separately
    current_space = sign_off_start - data_start_row
    
    if required_rows > current_space:
        # Need to move sign-off section down
        shift_amount = required_rows - current_space + 2  # +2 for spacing and Total row
        move_range_down(annexure_sheet, sign_off_start, sign_off_end, shift_amount)
        sign_off_start += shift_amount
    
    # Clear existing data between header and sign-off section
    for row in range(data_start_row, sign_off_start):
        for col in range(1, annexure_sheet.max_column + 1):
            cell = annexure_sheet.cell(row=row, column=col)
            cell.value = None
            if cell.has_style:
                new_style = copy(cell._style)
                cell._style = new_style
    
    # Get template row for formatting
    template_row = data_start_row
    template_cells = {col: annexure_sheet.cell(row=template_row, column=col) 
                     for col in range(1, annexure_sheet.max_column + 1)}
    
    # Define column mappings starting from column B
    column_mappings = {
        'B': 'S Loc Code',
        'C': 'Material Code',
        'D': 'Material Name',
        'E': 'UOM IN MT / No.',
        'F': 'Closing Balance - Net Weight (SAP)',
        'G': 'Physical Stock - Net Weight',
        'H': 'Stock under Fumigation',
        'I': 'Total Physical Stock',
        'J': 'Actual Shortage / Excess (-/+)',
        # 'K': 'Remark 1',
        # 'L': 'General Remark'
    }
    
    # Process data
    print("\nProcessing data rows...")
    data_without_total = processed_data[processed_data['Material Name'] != 'Total']
    print(f"Number of data rows to process: {len(data_without_total)}")
    physical_total = 0
    fumigation_total = 0
    
    for idx, (_, row_data) in enumerate(data_without_total.iterrows()):
        current_row = data_start_row + idx
        print(f"\nProcessing row {current_row}:")
        
        # Create a new row with template formatting
        for col in range(1, annexure_sheet.max_column + 1):
            target_cell = annexure_sheet.cell(row=current_row, column=col)
            template_cell = template_cells[col]
            copy_cell_format(template_cell, target_cell)
        
        # Fill in the data
        material_code = None
        for col_letter, col_name in column_mappings.items():
            try:
                col_idx = ord(col_letter) - ord('A') + 1
                cell = annexure_sheet.cell(row=current_row, column=col_idx)
                
                if col_name == 'Material Code':
                    material_code = row_data[col_name]
                    print(f"Material Code: {material_code}")
                
                # Handle special columns
                if col_name == 'Physical Stock - Net Weight' and material_code:
                    value = physical_stock_dict.get(material_code)
                    if value is None or not isinstance(value, (int, float)) or value == 0:
                        cell.value = ''
                    else:
                        cell.value = value
                        physical_total = round(physical_total + value, 2)
                    print(f"Physical Stock: {cell.value}")
                elif col_name == 'Stock under Fumigation' and material_code:
                    value = fumigation_stock_dict.get(material_code)
                    if value is None or not isinstance(value, (int, float)) or value == 0:
                        cell.value = ''
                    else:
                        cell.value = value
                        fumigation_total = round(fumigation_total + value, 2)
                    print(f"Fumigation Stock: {cell.value}")
                elif col_name == 'Total Physical Stock' and material_code:
                    # Set Excel formula for this row
                    cell.value = f"=G{current_row}+H{current_row}"
                elif col_name == 'Actual Shortage / Excess (-/+)' and material_code:
                    # Set Excel formula for this row
                    cell.value = f"=I{current_row}-F{current_row}"
                else:
                    cell.value = row_data[col_name]
            except Exception as e:
                print(f"Error processing cell {col_letter}{current_row}: {str(e)}")
                continue
    
    # Add Total row
    print("\nAdding Total row...")
    total_row = data_start_row + len(data_without_total)
    print(f"Total row will be at: {total_row}")
    print(f"Final Physical Total: {physical_total}")
    print(f"Final Fumigation Total: {fumigation_total}")
    
    for col in range(1, annexure_sheet.max_column + 1):
        cell = annexure_sheet.cell(row=total_row, column=col)
        if total_row_info and col <= len(total_row_info):
            # Apply formatting from existing Total row
            _, style = total_row_info[col-1]
            if style:
                cell._style = style
        
        # Set values for specific columns
        if col == 4:  # Column D (Material Name)
            cell.value = 'Total'
        elif col == 6:  # Column F (Closing Balance)
            # Calculate the sum from the data we just inserted
            start_col = ord('F') - ord('A') + 1
            start_row = data_start_row
            end_row = total_row - 1
            cell.value = f'=ROUND(SUBTOTAL(9,{chr(64 + start_col)}{start_row}:{chr(64 + start_col)}{end_row}), 2)'
        elif col == 7:  # Column G (Physical Stock)
            cell.value = physical_total
        elif col == 8:  # Column H (Fumigation Stock)
            cell.value = fumigation_total
        elif col == 9:  # Column I (Total Physical Stock)
            # Add vertical sum formula for Total Physical Stock
            start_row = data_start_row
            end_row = total_row - 1
            cell.value = f'=ROUND(SUM(I{start_row}:I{end_row}), 2)'
        elif col == 10:  # Column J (Actual Shortage/Excess)
            # Add vertical sum formula for Actual Shortage / Excess (-/+)
            start_row = data_start_row
            end_row = total_row - 1
            cell.value = f'=ROUND(SUM(J{start_row}:J{end_row}), 2)'
    
    print("\nSaving workbook...")
    wb.save(format_file_path)
    print("Update complete!")

def process_raw_material(format_file_path, master_data):
    processed_data = process_mb52_stock(format_file_path)
    physical_stock, fumigation_stock = process_rm_stack_wise(format_file_path)
    # Update processed_data with physical and fumigation stock
    for idx, row in processed_data.iterrows():
        material_code = row['Material Code']
        if pd.notna(material_code):
            processed_data.at[idx, 'Physical Stock - Net Weight'] = physical_stock.get(material_code, '')
            processed_data.at[idx, 'Stock under Fumigation'] = fumigation_stock.get(material_code, '')
    update_annexure_sheet(format_file_path, processed_data, master_data)

def main():
    pass

if __name__ == "__main__":
    main()
