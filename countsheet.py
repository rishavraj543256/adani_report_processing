import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

def find_signoff_section(sheet):
    """Find the start and end row of sign-off section"""
    start_row = None
    end_row = None
    
    # Find where the sign-off section starts by looking for "S Loc Incharge"
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and "S Loc Incharge" in str(cell_value):
                start_row = row
                # Find the end of sign-off section (first empty row after start)
                for end_r in range(start_row + 1, sheet.max_row + 1):
                    is_empty = True
                    for c in range(1, sheet.max_column + 1):
                        if sheet.cell(row=end_r, column=c).value:
                            is_empty = False
                            break
                    if is_empty:
                        end_row = end_r - 1
                        break
                if not end_row:  # If no empty row found
                    end_row = sheet.max_row
                return start_row, end_row
    
    return None, None

def move_range_down(sheet, start_row, end_row, shift_amount):
    """Move a range of cells down by shift_amount rows"""
    if start_row is None or end_row is None:
        return
        
    # Move from bottom to top to avoid overwriting
    for row in range(end_row, start_row - 1, -1):
        for col in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=row, column=col)
            target_cell = sheet.cell(row=row + shift_amount, column=col)
            
            # Copy value and style
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            
            # Clear source cell
            source_cell.value = None

def copy_cell_format(source_cell, target_cell):
    """Copy cell format from source to target"""
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)

def get_column_index_by_header(headers, header_name):
    """Get the column index (1-based) for a given header name"""
    try:
        return headers.index(header_name) + 1
    except ValueError:
        return None

def process_count_sheet(input_file, output_file):
    # Read the input file
    input_df = pd.read_excel(input_file)
    
    # Load the template workbook
    workbook = openpyxl.load_workbook(output_file)
    count_sheet = workbook['Count Sheet']
    
    # Find the sign-off section
    signoff_start, signoff_end = find_signoff_section(count_sheet)
    
    # Calculate required rows and move sign-off section if needed
    data_rows = len(input_df)
    if signoff_start is not None:
        current_data_space = signoff_start - 2  # Subtract header row
        if data_rows > current_data_space:
            shift_amount = data_rows - current_data_space + 2  # Add some buffer
            move_range_down(count_sheet, signoff_start, signoff_end, shift_amount)
    
    # Get headers from the count sheet (row 1)
    headers = []
    for col in range(1, count_sheet.max_column + 1):
        cell = count_sheet.cell(row=1, column=col)
        if cell.value:
            headers.append(cell.value)
    
    # Get column indices for special columns
    diff_col = get_column_index_by_header(headers, "Diff")
    gross_qty_col = get_column_index_by_header(headers, "Gross QTY")
    book_stock_col = get_column_index_by_header(headers, "Item QTY As Per book Stock")
    
    # Store reference cells from the second row
    reference_cells = {}
    for col in range(1, len(headers) + 1):
        reference_cells[col] = count_sheet.cell(row=2, column=col)
    
    # Start writing from row 2 (after headers)
    current_row = 2
    
    # Write data row by row
    for _, row in input_df.iterrows():
        for col, header in enumerate(headers, start=1):
            target_cell = count_sheet.cell(row=current_row, column=col)
            
            # Handle Diff column specially
            if col == diff_col:
                # Create Excel formula for Diff column with ABS for absolute value
                gross_qty_cell = get_column_letter(gross_qty_col) + str(current_row)
                book_stock_cell = get_column_letter(book_stock_col) + str(current_row)
                target_cell.value = f"=ABS({gross_qty_cell}-{book_stock_cell})"
            elif header in row:
                target_cell.value = row[header]
            
            # Copy format from reference cell
            if col in reference_cells:
                copy_cell_format(reference_cells[col], target_cell)
        
        current_row += 1
    
    # Save the workbook
    workbook.save(output_file)
