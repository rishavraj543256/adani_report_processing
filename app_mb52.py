from openpyxl import load_workbook
from copy import copy
from openpyxl.utils import get_column_letter

def process_mb52(format_file_path, mb52_input_file_path, s_loc_code):
    try:
        input_file = mb52_input_file_path
        output_file = format_file_path
        sheet_name = "Mb52- Stock Report"

        # Step 2: Open input workbook and read sheet
        wb_input = load_workbook(input_file,data_only=True)
        ws_input = wb_input.active

        # Step 3: Get headers from input sheet (assume headers are in row 3)
        header_row_input = 3
        headers_input = [cell.value for cell in ws_input[header_row_input]]

        # Step 4: Column mapping
        column_mapping = {
            "Plant": "Plant",
            "Material": "Material",
            "Material Description": "Material Description",
            "Include/ Exclude": "Exclusion/ Inclusion",
            "Name 1": "Name 1",
            "Storage Location": "Storage Location",
            "Category": "Category",
            "Descr. of Storage Loc.": "Descr. of Storage Loc.",
            "Base Unit of Measure": "Base Unit of Measure",
            "Unrestricted": "Unrestricted",
            "Quality Inspection": "Quality Inspection",
            "Blocked": "Blocked",
            "Returns": "Returns",
            "Transit and Transfer": "Transit and Transfer",
            "Restricted-Use Stock": "Restricted-Use Stock",
            "Special Stock": "Special Stock",
            "Total Qty": "Total Stock",
            "Value Unrestricted": "Value Unrestricted",
            "Value in QualInsp.": "Value in QualInsp.",
            "Value BlockedStock": "Value BlockedStock",
            "Value Rets Blocked": "Value Rets Blocked",
            "Val. in Trans./Tfr": "Val. in Trans./Tfr",
            "Value Restricted": "Value Restricted",
            "Total Value": "Total Value"
        }

        # Step 5: Filter rows by Storage Location
        data_rows = []
        for row in ws_input.iter_rows(min_row=header_row_input + 1, values_only=True):
            row_dict = dict(zip(headers_input, row))
            if str(row_dict.get("S Loc Code", "")).strip() == str(s_loc_code):
                data_rows.append(row_dict)

        if not data_rows:
            print("[WARNING] No data found for the given Storage Location.")
            return

        # Step 6: Open output workbook
        wb_output = load_workbook(output_file)
        ws_output = wb_output[sheet_name]

        # Step 7: Find sign-off section row
        sign_off_label = "S Loc Incharge (WMS Representative)"
        sign_off_row = None
        for row in ws_output.iter_rows(min_row=1, max_row=ws_output.max_row):
            for cell in row:
                if str(cell.value).strip() == sign_off_label:
                    sign_off_row = cell.row
                    break
            if sign_off_row:
                break
        if not sign_off_row:
            raise Exception("[ERROR] Sign-off label not found in output file!")

        # Step 8: Calculate available space and insert rows if needed
        header_start = 3
        insert_start = header_start + 1
        rows_to_write = len(data_rows)
        desired_gap = 4
        required_space = rows_to_write + desired_gap
        available_rows = sign_off_row - insert_start

        if required_space > available_rows:
            extra_rows = required_space - available_rows
            ws_output.insert_rows(idx=insert_start, amount=extra_rows)

        # Step 9: Copy formatting from row 3
        format_row_number = 3
        format_row_styles = []
        for col in range(1, ws_output.max_column + 1):
            cell = ws_output.cell(row=format_row_number, column=col)
            format_row_styles.append({
                "font": copy(cell.font),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
                "number_format": copy(cell.number_format),
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment)
            })

        # Apply formatting to new rows
        for i in range(rows_to_write):
            for col in range(1, ws_output.max_column + 1):
                tgt_cell = ws_output.cell(row=insert_start + i, column=col)
                style = format_row_styles[col - 1]
                tgt_cell.font = style["font"]
                tgt_cell.border = style["border"]
                tgt_cell.fill = style["fill"]
                tgt_cell.number_format = style["number_format"]
                tgt_cell.protection = style["protection"]
                tgt_cell.alignment = style["alignment"]

        # Step 10: Write data to output
        output_headers = [cell.value for cell in ws_output[2]]
        for i, row_data in enumerate(data_rows):
            # Calculate Total Stock
            total_stock = sum(float(row_data.get(col, 0) or 0) for col in [
                "Unrestricted",
                "Quality Inspection",
                "Blocked",
                "Returns",
                "Transit and Transfer",
                "Restricted-Use Stock",
                "Special Stock"
            ])
            # Calculate Total Value
            total_value = sum(float(row_data.get(col, 0) or 0) for col in [
                "Value Unrestricted",
                "Value in QualInsp.",
                "Value BlockedStock",
                "Value Rets Blocked",
                "Val. in Trans./Tfr",
                "Value Restricted"
            ])
            # Update the row data with calculated totals
            row_data["Total Qty"] = total_stock
            row_data["Total Value"] = total_value
            for j, out_header in enumerate(output_headers):
                in_header = next((k for k, v in column_mapping.items() if v == out_header), None)
                if in_header and in_header in row_data:
                    ws_output.cell(row=insert_start + i, column=j + 1, value=row_data[in_header])

        # Move the delete_rows before calculating totals
        ws_output.delete_rows(3)

        # Step 11: Add total row at the bottom
        total_row_index = insert_start + rows_to_write - 1
        ws_output.cell(row=total_row_index, column=2, value="TOTAL")
        output_headers = [cell.value for cell in ws_output[2]]
        total_columns = [
            "Unrestricted",
            "Quality Inspection",
            "Blocked",
            "Returns",
            "Transit and Transfer",
            "Restricted-Use Stock",
            "Special Stock",
            "Total Stock",
            "Value Unrestricted",
            "Value in QualInsp.",
            "Value BlockedStock",
            "Value Rets Blocked",
            "Val. in Trans./Tfr",
            "Value Restricted",
            "Total Value"
        ]
        for j, out_header in enumerate(output_headers):
            if out_header in total_columns:
                col_letter = get_column_letter(j + 1)
                start_row = insert_start
                end_row = insert_start + rows_to_write - 2
                if end_row >= start_row:
                    formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                    ws_output.cell(row=total_row_index, column=j + 1, value=formula)

        wb_output.save(output_file)
        print(f"[INFO] MB52 data processed and saved to {output_file}")
    except Exception as e:
        print(f"[ERROR] {e}")
