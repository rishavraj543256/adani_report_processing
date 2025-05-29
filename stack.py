import pandas as pd
import openpyxl
import numpy as np
from copy import copy
# from sign_off import write_value_below_label  # Not used in this context

def process_stack_data(input_file, output_file, master_data):
    try:
        output_sheet_name = 'RM- Stack wise'
        wb = openpyxl.load_workbook(output_file)
        ws = wb[output_sheet_name]

        # Write audit date and PSV quarter
        print("Filling header data...")
        ws['C2'] = master_data.get("Date of Audit")
        ws['C3'] = master_data.get("PSV Quarter")

        column_map = {
            'Stack No': 'Stack No. (With Stock)',
            'Material Code': 'Material Code',
            'Material Name': 'Material Name',
            'Normal Bag': 'Normal Bags',
            'Madeup Bag': 'Made up Bags',
            'Total Bags': 'Total Bags',
            'Gross QTY': 'Qty. In MT',
            'Bardana Weight': 'Bardana Weight',
        }

        def remove_empty_rows_in_section(ws, start_row, end_row, key_column):
            """
            Remove empty rows in a section while preserving the total row
            Args:
                ws: worksheet object
                start_row: starting row number (inclusive)
                end_row: ending row number (inclusive)
                key_column: column index to check for emptiness
            """
            rows_to_delete = []
            
            # First pass: identify rows to delete
            for row in range(start_row, end_row + 1):
                # Skip if it's a total row
                first_cell = str(ws.cell(row=row, column=1).value or '').strip().lower()
                if 'total' in first_cell:
                    continue
            
                # Check if row is empty
                is_empty = True
                for col in range(1, 9):  # Check first 8 columns
                    value = ws.cell(row=row, column=col).value
                    if value not in (None, '', '#DIV/0!', '-'):
                        is_empty = False
                        break
            
                if is_empty:
                    rows_to_delete.append(row)
        
            # Delete rows from bottom to top to avoid shifting issues
            for row in reversed(rows_to_delete):
                ws.delete_rows(row)
        
            return len(rows_to_delete)

        def print_cell_values(ws, start_row, end_row, col_range=(1, 5)):
            """Print cell values for debugging"""
            for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
                row_values = []
                for col in range(col_range[0], col_range[1] + 1):
                    value = ws.cell(row=row, column=col).value
                    row_values.append(f"{col}:'{value}'")
                print(f"Row {row}: {' | '.join(row_values)}")

        print("\n=== Starting Data Processing ===")
        print("\nReading input file:", input_file)
        input_df = pd.read_excel(input_file)
        print(f"Total rows in input file: {len(input_df)}")
        print("\nCleaning up input data...")
        valid_df = input_df[input_df['Stock Type'].notna()].copy()
        print(f"Rows after removing NaN Stock Type: {len(valid_df)}")
        valid_df.loc[:, 'Stock Type'] = valid_df['Stock Type'].str.lower().str.strip()
        valid_stock_types = ['general', 'fumigation']
        valid_df = valid_df[valid_df['Stock Type'].isin(valid_stock_types)]
        print("Valid stock types found:", valid_df['Stock Type'].unique())
        print(f"Final number of valid rows to process: {len(valid_df)}")
        stock_type_counts = valid_df['Stock Type'].value_counts()
        print("\nRows per stock type:")
        for stock_type, count in stock_type_counts.items():
            print(f"- {stock_type}: {count} rows")
        print("\nLocating sections in output template...")
        print(f"Total rows in sheet: {ws.max_row}")
        print("\nChecking rows 30-35 for fumigation section:")
        print_cell_values(ws, 30, 35)
        general_header_row = None
        fumigation_label_row = None
        total_stock_row = None
        name_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, 10):
                cell_value = str(ws.cell(row=row, column=col).value or '').strip()
                if cell_value == 'Stack No. (With Stock)':
                    general_header_row = row
                    print(f"Found General section header at row {row}, column {col}")
                    break
            if general_header_row:
                break
        if not general_header_row:
            print("Error: Could not find General section header 'Stack No. (With Stock)'")
            return
        for row in range(general_header_row + 1, ws.max_row + 1):
            col1_value = str(ws.cell(row=row, column=1).value or '').strip()
            if col1_value == 'Total Stock':
                total_stock_row = row
                print(f"Found Total Stock row at row {row}")
            for col in range(1, 5):
                cell_value = str(ws.cell(row=row, column=col).value or '').strip()
                if cell_value.lower() == 'stock under fumigation':
                    fumigation_label_row = row
                    print(f"Found Fumigation section at row {row}, column {col}")
                    break
            if col1_value == 'Name':
                name_row = row
                print(f"Found Name row at row {row}")
        if not fumigation_label_row:
            print("\nTrying broader search for Fumigation section...")
            for row in range(1, ws.max_row + 1):
                for col in range(1, 5):
                    cell_value = str(ws.cell(row=row, column=col).value or '').strip()
                    if 'fumigation' in cell_value.lower():
                        fumigation_label_row = row
                        print(f"Found Fumigation section using broader search at row {row}, column {col}")
                        break
                if fumigation_label_row:
                    break
        if not fumigation_label_row:
            print("Error: Could not find Fumigation section.")
            print("Creating a simplified version with just the General section.")
        else:
            print(f"Fumigation label row is at {fumigation_label_row}")
        fumigation_header_row = fumigation_label_row + 1 if fumigation_label_row else None
        if fumigation_header_row:
            print(f"Fumigation header row is at {fumigation_header_row}")
            print("\nChecking fumigation header row:")
            print_cell_values(ws, fumigation_header_row, fumigation_header_row)
        print("\nMapping columns...")
        general_col_idx = {}
        # for cell in ws[general_header_row]:
        #     if cell.value:
        #         general_col_idx[str(cell.value).strip()] = cell.column
        for cell in ws[general_header_row]:
            if cell.value:
                header_text = str(cell.value).strip().replace('\xa0', ' ')  # replaces non-breaking spaces
                general_col_idx[header_text] = cell.column

        print("Normalized header columns:", list(general_col_idx.keys()))

        print(f"Found {len(general_col_idx)} columns in General section")
        print("General column indices:", general_col_idx)
        print("Header row values and columns:")
        for cell in ws[general_header_row]:
            print(f"Column {cell.column}: '{cell.value}'")
        fumigation_col_idx = {}
        if fumigation_header_row:
            for cell in ws[fumigation_header_row]:
                if cell.value:
                    fumigation_col_idx[str(cell.value).strip()] = cell.column
            print(f"Found {len(fumigation_col_idx)} columns in Fumigation section")
        else:
            fumigation_col_idx = general_col_idx

        # --- Find sign-off section using 'S Loc Incharge' label (robust, like app_mb52.py) ---
        sign_off_label = "S Loc Incharge (WMS Representative)"
        signoff_start_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if str(cell.value).strip() == sign_off_label:
                    signoff_start_row = cell.row
                    break
            if signoff_start_row:
                break
        if not signoff_start_row:
            raise Exception("Could not find the sign-off section (row with 'S Loc Incharge (WMS Representative)')!")

        # --- Robust dynamic row management for General and Fumigation sections ---
        # 1. Count data rows needed
        general_data = valid_df[valid_df['Stock Type'] == 'general']
        fumigation_data = valid_df[valid_df['Stock Type'] == 'fumigation']
        num_general_rows = len(general_data)
        num_fumigation_rows = len(fumigation_data)
        gap_between_sections = 2
        gap_before_signoff = 2

        # 2. Count existing rows in each section
        # General section: from general_header_row+1 to (fumigation_label_row or signoff_start_row) - gap_between_sections
        if fumigation_label_row:
            general_section_end = fumigation_label_row - gap_between_sections
        else:
            general_section_end = signoff_start_row - gap_before_signoff
        existing_general_rows = general_section_end - (general_header_row + 1) + 1
        # Fumigation section: from fumigation_header_row+1 to signoff_start_row - gap_before_signoff
        if fumigation_label_row:
            existing_fumigation_rows = signoff_start_row - gap_before_signoff - (fumigation_header_row + 1) + 1
        else:
            existing_fumigation_rows = 0

        # 3. Insert rows as needed for General section
        general_insert_start = general_header_row + 1
        if num_general_rows > existing_general_rows:
            rows_to_insert = num_general_rows - existing_general_rows
            print(f"Inserting {rows_to_insert} rows after General header at row {general_insert_start}")
            ws.insert_rows(idx=general_insert_start, amount=rows_to_insert)
            if fumigation_label_row:
                fumigation_label_row += rows_to_insert
                if fumigation_header_row:
                    fumigation_header_row += rows_to_insert
            if signoff_start_row > general_insert_start:
                signoff_start_row += rows_to_insert

        # 4. Insert rows as needed for Fumigation section
        if fumigation_label_row and num_fumigation_rows > existing_fumigation_rows:
            fumigation_insert_start = fumigation_header_row + 1
            rows_to_insert = num_fumigation_rows - existing_fumigation_rows
            print(f"Inserting {rows_to_insert} rows after Fumigation header at row {fumigation_insert_start}")
            ws.insert_rows(idx=fumigation_insert_start, amount=rows_to_insert)
            if signoff_start_row > fumigation_insert_start:
                signoff_start_row += rows_to_insert

        # 5. Ensure 2-row gap between General and Fumigation
        if fumigation_label_row:
            gap_start = general_header_row + 1 + num_general_rows
            actual_gap = fumigation_label_row - gap_start
            if actual_gap != gap_between_sections:
                gap_diff = gap_between_sections - actual_gap
                print(f"Adjusting gap between General and Fumigation by {gap_diff} rows at {gap_start}")
                if gap_diff > 0:
                    ws.insert_rows(idx=gap_start, amount=gap_diff)
                    fumigation_label_row += gap_diff
                    if fumigation_header_row:
                        fumigation_header_row += gap_diff
                    if signoff_start_row > gap_start:
                        signoff_start_row += gap_diff
                elif gap_diff < 0:
                    ws.delete_rows(idx=gap_start, amount=abs(gap_diff))
                    fumigation_label_row += gap_diff
                    if fumigation_header_row:
                        fumigation_header_row += gap_diff
                    if signoff_start_row > gap_start:
                        signoff_start_row += gap_diff

        # Always copy formatting from Row 2, Column C (cell C2) for all data rows
        format_cell = ws.cell(row=2, column=3)  # C2
        print(f"Template cell C2: font={format_cell.font}, border={format_cell.border}, fill={format_cell.fill}, number_format={format_cell.number_format}")
        # Format General section rows (data only, not header)
        print(f"Formatting General section data rows {general_header_row + 1} to {general_header_row + num_general_rows} (columns 1 to {ws.max_column}) using C2 as template")
        for row in range(general_header_row + 1, general_header_row + num_general_rows + 1):
            for col in range(1, ws.max_column + 1):
                tgt_cell = ws.cell(row=row, column=col)
                tgt_cell.value = None
                tgt_cell.font = copy(format_cell.font)
                tgt_cell.border = copy(format_cell.border)
                tgt_cell.fill = copy(format_cell.fill)
                tgt_cell.number_format = format_cell.number_format
                tgt_cell.protection = copy(format_cell.protection)
                tgt_cell.alignment = copy(format_cell.alignment)
                print(f"Formatted General row {row}, col {col} with C2 style")
        # Format Fumigation section rows (data only, not header)
        if fumigation_label_row and num_fumigation_rows > 0:
            print(f"Formatting Fumigation data rows {fumigation_header_row + 1} to {fumigation_header_row + num_fumigation_rows} (columns 1 to {ws.max_column}) using C2 as template")
            for row in range(fumigation_header_row + 1, fumigation_header_row + num_fumigation_rows + 1):
                for col in range(1, ws.max_column + 1):
                    tgt_cell = ws.cell(row=row, column=col)
                    tgt_cell.value = None
                    tgt_cell.font = copy(format_cell.font)
                    tgt_cell.border = copy(format_cell.border)
                    tgt_cell.fill = copy(format_cell.fill)
                    tgt_cell.number_format = format_cell.number_format
                    tgt_cell.protection = copy(format_cell.protection)
                    tgt_cell.alignment = copy(format_cell.alignment)
                    print(f"Formatted Fumigation row {row}, col {col} with C2 style")
        # Write General section data
        row_ptr = general_header_row + 1
        print("General column indices:", general_col_idx)
        general_totals = {col: 0 for col in ['Normal Bags', 'Made up Bags', 'Total Bags', 'Gross QTY', 'Net Weight']}
        print("Available columns:", list(general_data.columns))
        # Find the Net Weight column index once
        net_weight_col = find_column_by_label(ws, general_header_row, "Net Weight")
        kgs_per_bag_col = find_column_by_label(ws, general_header_row, "Kgs per Bag*")
        if not net_weight_col:
            print("ERROR: Could not find 'Net Weight' column in header row!")
        if not kgs_per_bag_col:
            print("ERROR: Could not find 'Kgs per Bag*' column in header row!")
        else:
            for _, row in general_data.iterrows():
                gross_qty = row.get('Gross QTY', 0) or 0
                bardana_weight = row.get('Bardana Weight', 0) or 0
                category = str(master_data.get('Category', '')).strip().lower()

                # Calculate Bardana Weight as per category
                if category == 'wheat':
                    bardana_weight = gross_qty * (row['Bardana Weight'] / 100)
                elif category in ['paddy/rice', 'paddy', 'rice']:
                    total_bags = row['Normal Bag'] + row['Madeup Bag']
                    bardana_weight = total_bags * row['Bardana Weight'] / 1000

                net_weight = gross_qty - bardana_weight

                # Write all columns except 'Total Bags' as before
                for in_col, out_col in column_map.items():
                    if in_col in row and out_col in general_col_idx and out_col != 'Total Bags':
                        value = row[in_col]
                        if out_col == 'Bardana Weight':
                            value = round(bardana_weight, 2)
                        ws.cell(row=row_ptr, column=general_col_idx[out_col], value=value)
                        print(f"WROTE General: ws.cell(row={row_ptr}, col={general_col_idx[out_col]}) = {value} (for column '{out_col}')")
                # Write Excel formula for Total Bags
                if 'Normal Bags' in general_col_idx and 'Made up Bags' in general_col_idx and 'Total Bags' in general_col_idx:
                    normal_col = general_col_idx['Normal Bags']
                    madeup_col = general_col_idx['Made up Bags']
                    total_col = general_col_idx['Total Bags']
                    normal_cell = f"{openpyxl.utils.get_column_letter(normal_col)}{row_ptr}"
                    madeup_cell = f"{openpyxl.utils.get_column_letter(madeup_col)}{row_ptr}"
                    formula = f"={normal_cell}+{madeup_cell}"
                    ws.cell(row=row_ptr, column=total_col, value=formula)
                    print(f"WROTE General: ws.cell(row={row_ptr}, col={total_col}) = {formula} (for column 'Total Bags')")
                # Write Net Weight (Gross QTY - Bardana Weight)
                ws.cell(row=row_ptr, column=net_weight_col, value=round(net_weight, 2))
                print(f"WROTE Net Weight: ws.cell(row={row_ptr}, col={net_weight_col}) = {round(net_weight, 2)}")
                # Write Kgs per Bag formula
                if kgs_per_bag_col and 'Qty. In MT' in general_col_idx and 'Normal Bags' in general_col_idx:
                    qty_col = general_col_idx['Qty. In MT']
                    normal_bags_col = general_col_idx['Normal Bags']
                    qty_cell = f"{openpyxl.utils.get_column_letter(qty_col)}{row_ptr}"
                    normal_bags_cell = f"{openpyxl.utils.get_column_letter(normal_bags_col)}{row_ptr}"
                    formula = f"=IF({normal_bags_cell}=0,0,{qty_cell}*1000/{normal_bags_cell})"
                    ws.cell(row=row_ptr, column=kgs_per_bag_col, value=formula)
                    print(f"WROTE Kgs per Bag formula: ws.cell(row={row_ptr}, col={kgs_per_bag_col}) = {formula}")
                # Accumulate totals
                for col in general_totals:
                    if col in general_col_idx:
                        cell_value = ws.cell(row=row_ptr, column=general_col_idx[col]).value or 0
                        try:
                            general_totals[col] = round(general_totals.get(col, 0) + float(cell_value), 2)
                        except Exception:
                            pass
                print(f"Row {row_ptr}: Gross QTY = {gross_qty}, Bardana Weight = {bardana_weight}, Net Weight = {net_weight}")
                row_ptr += 1
        # Add Total row for General section
        print(f"Adding Total row for General section at row {row_ptr}")
        ws.cell(row=row_ptr, column=general_col_idx['Stack No. (With Stock)'], value='Total')
        for col in general_totals:
            if col in general_col_idx:
                value = general_totals[col]
                if col == 'Net Weight':
                    value = round(value, 2)
                ws.cell(row=row_ptr, column=general_col_idx[col], value=value)
        # Add Excel formula for Qty. In MT total (vertical sum)
        if 'Qty. In MT' in general_col_idx:
            qty_col = general_col_idx['Qty. In MT']
            data_start = general_header_row + 1
            data_end = row_ptr - 1
            if data_end >= data_start:
                formula = f"=SUM({openpyxl.utils.get_column_letter(qty_col)}{data_start}:{openpyxl.utils.get_column_letter(qty_col)}{data_end})"
                ws.cell(row=row_ptr, column=qty_col, value=formula)
        # Add Excel formula for Total Bags total (vertical sum)
        if 'Total Bags' in general_col_idx:
            total_bags_col = general_col_idx['Total Bags']
            data_start = general_header_row + 1
            data_end = row_ptr - 1
            if data_end >= data_start:
                formula = f"=SUM({openpyxl.utils.get_column_letter(total_bags_col)}{data_start}:{openpyxl.utils.get_column_letter(total_bags_col)}{data_end})"
                ws.cell(row=row_ptr, column=total_bags_col, value=formula)

        # --- Ensure 2-row gap after General Total before Fumigation section ---
        if fumigation_label_row:
            gap_start = row_ptr + 1
            actual_gap = fumigation_label_row - gap_start
            gap_needed = 2
            if actual_gap != gap_needed:
                gap_diff = gap_needed - actual_gap
                print(f"Adjusting gap after General Total before Fumigation by {gap_diff} rows at {gap_start}")
                if gap_diff > 0:
                    ws.insert_rows(idx=gap_start, amount=gap_diff)
                    fumigation_label_row += gap_diff
                    if fumigation_header_row:
                        fumigation_header_row += gap_diff
                    if signoff_start_row > gap_start:
                        signoff_start_row += gap_diff
                elif gap_diff < 0:
                    ws.delete_rows(idx=gap_start, amount=abs(gap_diff))
                    fumigation_label_row += gap_diff
                    if fumigation_header_row:
                        fumigation_header_row += gap_diff
                    if signoff_start_row > gap_start:
                        signoff_start_row += gap_diff

        # Write Fumigation section header and data (if present)
        if fumigation_label_row and num_fumigation_rows > 0:
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=fumigation_header_row, column=col)
                tgt_cell = ws.cell(row=fumigation_header_row, column=col)
                tgt_cell.value = src_cell.value
            # Write data rows
            row_ptr = fumigation_header_row + 1
            fumigation_totals = {col: 0 for col in ['Normal Bags', 'Made up Bags', 'Total Bags', 'Gross QTY', 'Net Weight']}
            for _, row in fumigation_data.iterrows():
                gross_qty = row.get('Gross QTY', 0) or 0
                bardana_weight = row.get('Bardana Weight', 0) or 0
                category = str(master_data.get('Category', '')).strip().lower()

                # Calculate Bardana Weight as per category
                if category == 'wheat':
                    bardana_weight = gross_qty * (row['Bardana Weight'] / 100)
                elif category in ['paddy/rice', 'paddy', 'rice']:
                    total_bags = row['Normal Bag'] + row['Madeup Bag']
                    bardana_weight = total_bags * row['Bardana Weight'] / 1000

                net_weight = gross_qty - bardana_weight

                # Write all columns except 'Total Bags' as before
                for in_col, out_col in column_map.items():
                    if in_col in row and out_col in fumigation_col_idx and out_col != 'Total Bags':
                        value = row[in_col]
                        if out_col == 'Bardana Weight':
                            value = round(bardana_weight, 2)
                        ws.cell(row=row_ptr, column=fumigation_col_idx[out_col], value=value)
                        print(f"WROTE Fumigation: ws.cell(row={row_ptr}, col={fumigation_col_idx[out_col]}) = {value} (for column '{out_col}')")
                # Write Excel formula for Total Bags
                if 'Normal Bags' in fumigation_col_idx and 'Made up Bags' in fumigation_col_idx and 'Total Bags' in fumigation_col_idx:
                    normal_col = fumigation_col_idx['Normal Bags']
                    madeup_col = fumigation_col_idx['Made up Bags']
                    total_col = fumigation_col_idx['Total Bags']
                    normal_cell = f"{openpyxl.utils.get_column_letter(normal_col)}{row_ptr}"
                    madeup_cell = f"{openpyxl.utils.get_column_letter(madeup_col)}{row_ptr}"
                    formula = f"={normal_cell}+{madeup_cell}"
                    ws.cell(row=row_ptr, column=total_col, value=formula)
                    print(f"WROTE Fumigation: ws.cell(row={row_ptr}, col={total_col}) = {formula} (for column 'Total Bags')")
                # Write Net Weight (Gross QTY - Bardana Weight)
                ws.cell(row=row_ptr, column=net_weight_col, value=round(net_weight, 2))
                print(f"WROTE Net Weight: ws.cell(row={row_ptr}, col={net_weight_col}) = {round(net_weight, 2)}")
                # Write Kgs per Bag formula
                if kgs_per_bag_col and 'Qty. In MT' in fumigation_col_idx and 'Normal Bags' in fumigation_col_idx:
                    qty_col = fumigation_col_idx['Qty. In MT']
                    normal_bags_col = fumigation_col_idx['Normal Bags']
                    qty_cell = f"{openpyxl.utils.get_column_letter(qty_col)}{row_ptr}"
                    normal_bags_cell = f"{openpyxl.utils.get_column_letter(normal_bags_col)}{row_ptr}"
                    formula = f"=IF({normal_bags_cell}=0,0,{qty_cell}*1000/{normal_bags_cell})"
                    ws.cell(row=row_ptr, column=kgs_per_bag_col, value=formula)
                    print(f"WROTE Kgs per Bag formula: ws.cell(row={row_ptr}, col={kgs_per_bag_col}) = {formula}")
                # Accumulate totals
                for col in fumigation_totals:
                    if col in fumigation_col_idx:
                        cell_value = ws.cell(row=row_ptr, column=fumigation_col_idx[col]).value or 0
                        try:
                            fumigation_totals[col] = round(fumigation_totals.get(col, 0) + float(cell_value), 2)
                        except Exception:
                            pass
                print(f"Row {row_ptr}: Gross QTY = {gross_qty}, Bardana Weight = {bardana_weight}, Net Weight = {net_weight}")
                row_ptr += 1
            # Add Total row for Fumigation section
            print(f"Adding Total row for Fumigation section at row {row_ptr}")
            ws.cell(row=row_ptr, column=fumigation_col_idx['Stack No. (With Stock)'], value='Total')
            for col in fumigation_totals:
                if col in fumigation_col_idx:
                    value = fumigation_totals[col]
                    if col == 'Net Weight':
                        value = round(value, 2)
                    ws.cell(row=row_ptr, column=fumigation_col_idx[col], value=value)
            # Add Excel formula for Qty. In MT total (vertical sum)
            if 'Qty. In MT' in fumigation_col_idx:
                qty_col = fumigation_col_idx['Qty. In MT']
                data_start = fumigation_header_row + 1
                data_end = row_ptr - 1
                if data_end >= data_start:
                    formula = f"=SUM({openpyxl.utils.get_column_letter(qty_col)}{data_start}:{openpyxl.utils.get_column_letter(qty_col)}{data_end})"
                    ws.cell(row=row_ptr, column=qty_col, value=formula)
            # Add Excel formula for Total Bags total (vertical sum)
            if 'Total Bags' in fumigation_col_idx:
                total_bags_col = fumigation_col_idx['Total Bags']
                data_start = fumigation_header_row + 1
                data_end = row_ptr - 1
                if data_end >= data_start:
                    formula = f"=SUM({openpyxl.utils.get_column_letter(total_bags_col)}{data_start}:{openpyxl.utils.get_column_letter(total_bags_col)}{data_end})"
                    ws.cell(row=row_ptr, column=total_bags_col, value=formula)

            # --- Ensure 2-row gap after Fumigation Total before sign-off section ---
            gap_start = row_ptr + 1
            actual_gap = signoff_start_row - gap_start
            gap_needed = 2
            if actual_gap != gap_needed:
                gap_diff = gap_needed - actual_gap
                print(f"Adjusting gap after Fumigation Total before sign-off by {gap_diff} rows at {gap_start}")
                if gap_diff > 0:
                    ws.insert_rows(idx=gap_start, amount=gap_diff)
                    if signoff_start_row > gap_start:
                        signoff_start_row += gap_diff
                elif gap_diff < 0:
                    ws.delete_rows(idx=gap_start, amount=abs(gap_diff))
                    if signoff_start_row > gap_start:
                        signoff_start_row += gap_diff

        wb.save(output_file)
        print(f"Stack data processed and saved to {output_file}")
    except Exception as e:
        print(f"Error processing stack data: {str(e)}")

def find_column_by_label(ws, header_row, label):
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip().lower() == label.strip().lower():
            return cell.column
    return None
