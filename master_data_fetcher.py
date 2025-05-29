import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from copy import copy

def format_current_date():
    """Format current date as '21st May' 25'"""
    today = datetime.today()
    day = today.day
    suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    return f"{day}{suffix} {today.strftime('%B')}' {today.strftime('%y')}"

def get_quarter_data(sheet):
    """Extract quarter data from the master sheet"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and 'Audit Quarter:' in str(cell.value):
                # Get the cell to the right of 'Quarter'
                quarter_value = sheet.cell(row=cell.row, column=cell.column+1).value
                if quarter_value:
                    return str(quarter_value).strip()
    return None

def process_auditor_names(auditor_names_str):
    """
    Process auditor names string based on different cases:
    Case 1: Single name -> Goes to Auditor Name 1
    Case 2: Two names separated by / -> Split between Auditor Name 1 and 2
    Case 3: Three names separated by / -> Split between Auditor Name 1, 2, and 3
    Case 4: Four names separated by / -> Split between all four Auditor Name fields
    """
    if not auditor_names_str or pd.isna(auditor_names_str):
        return {}
    
    # Clean and split the names
    names = [name.strip() for name in str(auditor_names_str).split('/')]
    
    # Create dictionary for auditor names
    auditor_data = {}
    for idx, name in enumerate(names, 1):
        if idx <= 4:  # Only process up to 4 auditors
            auditor_data[f"Auditor Name {idx}"] = name.strip()
    
    return auditor_data

def fetch_master_data(s_loc_code=None, category=None, master_file_path=None):
    """Fetch all required data from master sheet and return as a dictionary"""
    # Column mapping between format sheet labels and master sheet columns
    column_mapping = {
        "PSV Quarter": "Audit Quarter:",
        "Plant Code": "Plant",
        "S Loc Code": "S Loc",
        "S Loc Location": "City",
        "S Loc Address": "Address",
        "S Loc Incharge (WMS Representative)": "Contact Person",
        "WMS Service Provider":"WSP Agency",
        "Category": "Product Category"
    }

    try:
        if not master_file_path:
            return {
                "master_data": {},
                "auditor_data": {},
                "status": "error",
                "error_message": "Master file path is required"
            }

        # Load master data using pandas and openpyxl
        master_df = pd.read_excel(master_file_path, header=2)
        master_wb = load_workbook(master_file_path)
        master_sheet = master_wb.active

        # Convert S Loc to string for comparison
        master_df["S Loc"] = master_df["S Loc"].astype(str)
        
        # Filter data based on S Loc Code and Category if provided
        if s_loc_code:
            s_loc_code = str(s_loc_code).strip()  # Convert input to string and strip whitespace
            master_df = master_df[master_df["S Loc"].str.strip() == s_loc_code]
        if category:
            category = str(category).strip().lower()  # Convert to lowercase and strip whitespace
            master_df = master_df[master_df["Product Category"].str.strip().str.lower() == category]

        # Check if any data remains after filtering
        if master_df.empty:
            return {
                "master_data": {},
                "auditor_data": {},
                "status": "error",
                "error_message": f"No data found for S Loc Code: {s_loc_code} and Category: {category}"
            }

        # Initialize data dictionaries
        master_data = {}
        auditor_data = {}

        data_row = master_df.iloc[0]
        
        # Fill data based on column mapping
        for format_label, master_col in column_mapping.items():
            if master_col in master_df.columns:
                value = data_row[master_col]
                if pd.notna(value):
                    master_data[format_label] = str(value) if not isinstance(value, (int, float)) else value
                else:
                    master_data[format_label] = ""  # Set empty string for null/NaN values
            else:
                master_data[format_label] = ""  # Set empty string for missing columns

        # Add current date
        master_data["Date of Audit"] = format_current_date()
        
        # Add quarter value if found
        quarter_value = get_quarter_data(master_sheet)
        if quarter_value:
            master_data["PSV Quarter"] = quarter_value
        
        # Process auditor names if present
        if "Auditor's Name" in master_df.columns:
            auditor_names = data_row["Auditor's Name"]
            auditor_data = process_auditor_names(auditor_names)
            master_data.update(auditor_data)

        # Add Audit Firm (if present)
        if "Audit Firm" in master_df.columns and pd.notna(data_row["Audit Firm"]):
            master_data["Audit Firm"] = str(data_row["Audit Firm"])

        return {
            "master_data": master_data,
            "auditor_data": auditor_data,
            "status": "success"
        }

    except Exception as e:
        return {
            "master_data": {},
            "auditor_data": {},
            "status": "error",
            "error_message": str(e)
        }

def find_signoff_section(sheet):
    """Find the start and end row of sign-off section"""
    start_row = None
    end_row = None
    
    # Find where the sign-off section starts by looking for "S Loc Incharge"
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and "S Loc Incharge" in str(cell_value):
                start_row = row
                # Find the end of sign-off section (first empty row after start)
                for end_r in range(start_row + 1, sheet.max_row + 1):
                    is_empty = True
                    for c in range(1, sheet.max_column + 1):
                        if sheet.cell(row=end_r, column=c).value:
                            is_empty = False
                            break
                    if is_empty:
                        end_row = end_r - 1
                        break
                if not end_row:  # If no empty row found
                    end_row = sheet.max_row
                return start_row, end_row
    
    return None, None

def fill_signoff_section(sheet, master_data, auditor_data):
    """Fill the sign-off section with data from master sheet"""
    signoff_start, signoff_end = find_signoff_section(sheet)
    if not signoff_start:
        return

    # Map of fields to look for and their corresponding master data keys
    field_mapping = {
        "S Loc Incharge": "S Loc Incharge (WMS Representative)",
        "Auditor 1": "Auditor Name 1",
        "Auditor 2": "Auditor Name 2",
        "Auditor 3": "Auditor Name 3",
        "Auditor 4": "Auditor Name 4"
    }

    # Fill in the values
    for row in range(signoff_start, signoff_end + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = str(cell.value) if cell.value else ""

            # Check for each field in the mapping
            for field, master_key in field_mapping.items():
                if field in cell_value:
                    # Get the cell below (for the name)
                    name_cell = sheet.cell(row=row + 1, column=col)
                    
                    # Get value from master data or auditor data
                    if field.startswith("Auditor"):
                        value = auditor_data.get(master_key, "")
                    else:
                        value = master_data.get(master_key, "")
                    
                    name_cell.value = value

def update_all_sheets_signoff(master_data, auditor_data, format_file_path):
    try:
        wb = load_workbook(format_file_path)
        sheets_to_update = [
            "Annexure- Raw Material",
            "RM- Stack wise",
            "Annexure- Hygiene Obs",
            "Count Sheet",
            "Mb52- Stock Report"
        ]
        for sheet_name in sheets_to_update:
            if sheet_name in wb.sheetnames:
                print(f"Updating Sign Off section in {sheet_name}...")
                sheet = wb[sheet_name]
                fill_signoff_section(sheet, master_data, auditor_data)
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in {format_file_path}")
        wb.save(format_file_path)
        print("Successfully updated Sign Off sections in all sheets")
    except Exception as e:
        print(f"Error updating Sign Off sections: {str(e)}")

if __name__ == "__main__":
    print("\nPlease enter the following information:")
    s_loc_code = input("Enter S Loc Code: ").strip()
    category = input("Enter Category: ").strip()

    fetched_data = fetch_master_data(s_loc_code, category)

    # Print fetched data only when running this script directly
    print("\n=================== Data Fetched from Master Sheet ===================\n")
    print("\nMaster Data:")
    for key, value in fetched_data["master_data"].items():
        print(f"{key}: {value}")

    print("\nAuditor Data:")
    for key, value in fetched_data["auditor_data"].items():
        print(f"{key}: {value}")

    if fetched_data["status"] == "error":
        print(f"\nError: {fetched_data['error_message']}")
    print("\n=================== End of Fetched Data ==============================\n")

    # Update Sign Off sections only when running this script directly
    update_all_sheets_signoff(fetched_data["master_data"], fetched_data["auditor_data"], "format.xlsx") 