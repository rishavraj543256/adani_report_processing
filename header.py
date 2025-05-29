import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from master_data_fetcher import fetch_master_data

def format_current_date():
    """Format current date as '21st May' 25'"""
    today = datetime.today()
    day = today.day
    suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    return f"{day}{suffix} {today.strftime('%B')}' {today.strftime('%y')}"

def get_quarter_data(sheet):
    """Extract quarter data from the master sheet"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and 'Quarter' in str(cell.value):
                # Get the cell to the right of 'Quarter'
                quarter_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                if quarter_value:
                    return str(quarter_value).strip()
    return None

def find_sections(sheet):
    """
    Dynamically identify different sections in the format sheet
    Returns a dictionary of section info with their row ranges
    """
    sections = {}
    current_section = None
    section_headers = ["About the Location", "Auditor Details", "Sign Off"]
    last_content_row = 1
    
    # First pass: Find the last row with actual content
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        has_content = any(cell.value for cell in row)
        if has_content:
            last_content_row = row_idx
    
    # Second pass: Find sections
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        if row_idx > last_content_row:
            break
            
        for cell in row:
            if cell.value:
                cell_text = str(cell.value).strip()
                
                # If we found a section header
                if cell_text in section_headers:
                    # Close previous section if exists
                    if current_section:
                        sections[current_section]["end_row"] = row_idx - 1
                    
                    # Start new section
                    current_section = cell_text
                    sections[current_section] = {
                        "start_row": row_idx,
                        "end_row": last_content_row,
                        "column": cell.column
                    }
                    break
    
    # Adjust section end rows based on next section's start
    sorted_sections = sorted(sections.items(), key=lambda x: x[1]["start_row"])
    for i in range(len(sorted_sections) - 1):
        current_section = sorted_sections[i][0]
        next_section = sorted_sections[i + 1][0]
        sections[current_section]["end_row"] = sections[next_section]["start_row"] - 1
    
    return sections

def process_auditor_names(auditor_names_str):
    """
    Process auditor names string based on different cases:
    Case 1: Single name -> Goes to Auditor Name 1
    Case 2: Two names separated by / -> Split between Auditor Name 1 and 2
    Case 3: Three names separated by / -> Split between Auditor Name 1, 2, and 3
    Case 4: Four names separated by / -> Split between all four Auditor Name fields
    """
    if not auditor_names_str or pd.isna(auditor_names_str):
        return {}
    
    # Clean and split the names
    names = [name.strip() for name in str(auditor_names_str).split('/')]
    
    # Create dictionary for auditor names
    auditor_data = {}
    for idx, name in enumerate(names, 1):
        if idx <= 4:  # Only process up to 4 auditors
            auditor_data[f"Auditor Name {idx}"] = name.strip()
    
    return auditor_data

def find_name_row_in_signoff(sheet):
    """Find the row with 'Name' label in sign-off section"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == "Name":
                return cell.row, cell.column
    return None, None

def fill_signoff_section(sheet, master_data, auditor_names):
    """Fill the sign-off section with WMS Representative and auditor names"""
    name_row, name_col = find_name_row_in_signoff(sheet)
    if name_row:
        # Get the header row (one row above the 'Name' row)
        header_row = name_row - 1
        
        # Find columns for each field in the header row
        header_mapping = {}
        for cell in sheet[header_row]:
            value = str(cell.value).strip() if cell.value else ""
            if "S Loc Incharge (WMS Representative)" in value:
                header_mapping["wms_rep"] = cell.column
            elif "Auditor 1" in value:
                header_mapping["auditor1"] = cell.column
            elif "Auditor 2" in value:
                header_mapping["auditor2"] = cell.column
        
        # Fill WMS Representative if available
        if "wms_rep" in header_mapping:
            wms_rep = master_data.get("S Loc Incharge (WMS Representative)", "")
            if wms_rep and not pd.isna(wms_rep):
                sheet.cell(row=name_row, column=header_mapping["wms_rep"]).value = wms_rep
        
        # Fill auditor names if available
        auditor_list = list(auditor_names.values())
        if auditor_list:
            if "auditor1" in header_mapping and len(auditor_list) >= 1:
                sheet.cell(row=name_row, column=header_mapping["auditor1"]).value = auditor_list[0]
            if "auditor2" in header_mapping and len(auditor_list) >= 2:
                sheet.cell(row=name_row, column=header_mapping["auditor2"]).value = auditor_list[1]

def is_protected_cell(cell, sections):
    """Check if a cell should be protected from overwriting"""
    row = cell.row
    cell_value = str(cell.value).strip() if cell.value else ""
    
    # Only protect headers in the Sign Off section
    if "Sign Off" in sections:
        sign_off_section = sections["Sign Off"]
        if sign_off_section["start_row"] <= row <= sign_off_section["end_row"]:
            protected_values = ["Auditor 1", "Auditor 2", "S Loc Incharge (WMS Representative)"]
            if cell_value in protected_values:
                return True
    
    return False

def update_all_sheets_signoff(master_data, auditor_data, format_file_path):
    try:
        wb = load_workbook(format_file_path)
        sheets_to_update = [
            "Annexure- Raw Material",
            "RM- Stack wise",
            "Annexure- Hygiene Obs",
            "Count Sheet",
            "Mb52- Stock Report"
        ]
        for sheet_name in sheets_to_update:
            if sheet_name in wb.sheetnames:
                print(f"Updating Sign Off section in {sheet_name}...")
                sheet = wb[sheet_name]
                fill_signoff_section(sheet, master_data, auditor_data)
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in {format_file_path}")
        wb.save(format_file_path)
        print("Successfully updated Sign Off sections in all sheets")
    except Exception as e:
        print(f"Error updating Sign Off sections: {str(e)}")

def main(master_data=None, auditor_data=None, output_file_path=None, format_file_path=None):
    if not master_data or not auditor_data:
        print("❌ Error: Master data and auditor data are required")
        return

    if not output_file_path:
        print("❌ Error: Output file path is required")
        return

    if not format_file_path:
        print("❌ Error: Format file path is required")
        return

    try:
        # Load format workbook
        format_wb = load_workbook(format_file_path)
        format_sheet = format_wb["Header"]

        # DEBUG: Write a test value to A1
        #format_sheet['A1'] = "TEST VALUE - If you see this, file is being updated!"

        # Get section information
        sections = find_sections(format_sheet)

        # Fill data in format sheet
        for row in format_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Skip protected cells and AWL Executive
                    if is_protected_cell(cell, sections) or str(cell.value).strip() == "AWL Executive":
                        continue
                        
                    label = str(cell.value).strip()
                    if label in master_data:
                        right_cell = cell.offset(column=1)
                        right_cell.value = master_data[label]

        # Fill sign-off section with WMS Representative and auditor names only
        fill_signoff_section(format_sheet, master_data, auditor_data)

        # Save updated file
        try:
            format_wb.save(output_file_path)
            print(f"✅ Data mapped and filled successfully! Saved to: {output_file_path}")
        except Exception as e:
            print(f"❌ Error saving file: {e}")

        # Update sign-off in all relevant sheets
        update_all_sheets_signoff(master_data, auditor_data, format_file_path)

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        raise e

if __name__ == "__main__":
    main()